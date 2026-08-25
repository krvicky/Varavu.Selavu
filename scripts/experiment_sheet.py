"""Categorization experiment: merchant grouping -> Excel review sheet -> report.

    python scripts/experiment_sheet.py merchants   # print merchant groups (input for suggestions.csv)
    python scripts/experiment_sheet.py build       # transactions.csv + suggestions.csv -> categorization_review.xlsx
    python scripts/experiment_sheet.py apply       # edited xlsx -> experiment_report.md
    python scripts/experiment_sheet.py migrate     # rewrite category paths in an edited xlsx after a taxonomy change (keeps your edits)

All data files live in experiments/ (gitignored: real financial data).

History: the Kotak (May-Jul 2026) and HDFC Diners (May-Jul 2026) runs on 2026-08-17 produced the
reviewed decisions that seeded app.DEFAULT_RULES and SUBCATEGORIES. To onboard a new bank: parse it
with scripts/parse_statements.py --out experiments/transactions_<name>.csv, write
experiments/suggestions_<name>.csv, then `--dataset <name> build` / review / `apply`, and fold the
report's rules into DEFAULT_RULES by hand (tests/test_rules_coverage_local.py guards coverage).
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
EXP = ROOT / "experiments"
TXN_CSV = EXP / "transactions.csv"
SUGGESTIONS_CSV = EXP / "suggestions.csv"
XLSX = EXP / "categorization_review.xlsx"
REPORT = EXP / "experiment_report.md"


def use_dataset(name: str | None) -> None:
    """--dataset diners -> transactions_diners.csv / suggestions_diners.csv / categorization_review_diners.xlsx / experiment_report_diners.md"""
    global TXN_CSV, SUGGESTIONS_CSV, XLSX, REPORT
    if not name:
        return
    TXN_CSV = EXP / f"transactions_{name}.csv"
    SUGGESTIONS_CSV = EXP / f"suggestions_{name}.csv"
    XLSX = EXP / f"categorization_review_{name}.xlsx"
    REPORT = EXP / f"experiment_report_{name}.md"

SEP = " › "
FLOW_CHOICES = ["spend", "income", "refund", "transfer", "card_payment", "reversal", "fee", "exclude"]

# ---------------------------------------------------------------- merchant key


def _clean(s: str) -> str:
    s = s.upper().replace("&", " AND ")
    s = re.sub(r"[^A-Z0-9 ]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _upi_merchant(desc: str) -> str:
    """Merchant field of a 'UPI/<merchant>/<...>' description."""
    parts = desc.split("/")
    name = parts[1] if len(parts) > 1 else desc
    # "UPI/M/S.KLEVER K9/..." -> the merchant spans two slash-fields
    if name.strip().upper() in ("M", "MS") and len(parts) > 2:
        name = parts[2]
    name = re.sub(r"^\s*(M/?)?S\.\s*", "", name, flags=re.I)
    name = re.sub(r"^\s*(MR|MRS|MS|DR)\.?\s+", "", name, flags=re.I)
    name = _clean(name)
    name = re.sub(r"\b(LIMITED|LTD|PRIVATE|PRIVAT|PRIVA|PVT)\b", "", name)
    name = re.sub(r"\d+$", "", name.strip())
    return re.sub(r"\s+", " ", name).strip()


def merchant_key(desc: str) -> str:
    """Stable, human-readable key that groups the same counterparty together."""
    d = desc.strip()
    u = d.upper()
    if "UPI/" in u and (u.startswith("REV-") or u.startswith("REV ")):
        d = "UPI/" + d.split("UPI/", 1)[1]
        u = d.upper()
    if u.startswith("UPI_CRADJ"):
        return "UPI CREDIT ADJUSTMENT"
    if u.startswith("UPI/"):
        name = _upi_merchant(d)
        # the bank truncates names to ~14-15 chars; collapse to a stable prefix
        compact = name.replace(" ", "")
        return compact[:11]
    m = re.match(r"^(PCI|PCD)/\d+/(.+?)/", d)
    if m:
        return _clean(m.group(2).split("*")[0])
    if u.startswith("ATL/"):
        return "ATM WITHDRAWAL"
    if u.startswith("NACH-MUT-DR-"):
        return _clean(u[len("NACH-MUT-DR-"):].split("-")[0])[:20]
    if re.match(r"^NACH-(ECS|10)-CR-", u):
        return "NACH CREDIT (DIVIDEND/INTEREST)"
    if u.startswith("FD PREMAT PROCEEDS"):
        return "FD PREMATURE PROCEEDS"
    if u.startswith("SWEEP TRF FROM"):
        return "SWEEP IN FROM FD"
    if u.startswith("SWEEP TRANSFER TO"):
        return "SWEEP OUT TO FD"
    if "DCC FEE" in u:
        return "DCC FEE"
    if u.startswith("CHRG:") or u.startswith("REM CHRG:"):
        return _clean(re.sub(r"FOR .*$", "", u.split(":", 1)[1]))
    if u.startswith("NEFT") and "EMPLOYER PAYROLL" in u:
        return "EMPLOYER PAYROLL SALARY"
    # Axis incoming NEFT: "NEFT/<reference>/<counterparty>/<bank>/<remarks>" --
    # the reference changes every month, so key on the counterparty field.
    m = re.match(r"^NEFT/[A-Z0-9]+/(.+?)/", u)
    if m:
        return ("NEFT " + _clean(m.group(1))[:20]).strip()
    if "ITDTAX REFUND" in u or "IT REFUND" in u:
        return "INCOME TAX REFUND"
    if u.startswith("SENTIMPS"):
        return "IMPS " + _clean(re.sub(r"^SENTIMPS\d+", "", u).split("/")[0])
    if "SENT NEFT" in u:
        after = u.split("SENT NEFT", 1)[1]
        toks = [t for t in re.split(r"[ /]", after) if t and not re.search(r"\d", t)]
        return "NEFT " + " ".join(toks[:2])
    if u.startswith("KR, TT") and "INDMONE" in u:
        return "INDMONEY USD REMITTANCE"
    if u.startswith("INT.PD") or ":INT.PD:" in u:
        return "SAVINGS INTEREST"
    if u.startswith("MB:") or u.startswith("IB:"):
        return _clean(u[3:])
    if u.startswith("PG OSHDFCCC"):
        return "HDFC CREDIT CARD PAYMENT (PG)"
    # --- credit-card statement shapes (HDFC Diners): "MERCHANTCITY", fee lines with (Ref# ...)
    if "ONLINE PYMT RECD" in u or "PAYMENT RECEIVED" in u:
        return "CARD PAYMENT RECEIVED"
    if u.startswith("IGST"):
        return "IGST ON FEES"
    if "FCY MARKUP" in u:
        return "FCY MARKUP FEE"
    if "DCC TRANSACTION" in u:
        return "DCC MARKUP FEE"
    if u.startswith("PETRO SURCHARGE"):
        return "PETRO SURCHARGE WAIVER"
    return _card_merchant(d)


_CITIES = ["BANGALORE", "BENGALURU", "MUMBAI", "GURGAON", "GURUGRAM", "PUNE", "CHENNAI", "NOIDA", "KODAGU", "CYPRUS", "NICOSIA",
           "DELHI", "NEW DELHI", "HYDERABAD", "KOLKATA", "MYSORE", "MYSURU", "COORG", "OOTY", "GOA", "SINGAPORE", "LONDON", "DUBLIN"]
_CITY_RE = re.compile(r"(?:" + "|".join(sorted(_CITIES, key=len, reverse=True)) + r")\s*$", re.I)


def _card_merchant(desc: str) -> str:
    """'AMAZON SELLER SERVICESBANGALORE' -> 'AMAZON SELLER SERVICES'; 'CANVA* I04887-...' -> 'CANVA';
    'FIGMA4158905404' -> 'FIGMA'; 'upg*paymentico.comNicosia' -> 'UPG PAYMENTICO COM'."""
    d = re.sub(r"\(Ref#.*?\)", "", desc)
    d = d.split("*")[0] if "*" in d and len(d.split("*")[0].strip()) >= 4 else d.replace("*", " ")
    d = _CITY_RE.sub("", d.strip())
    d = _clean(d)
    d = re.sub(r"\b(PRIVATE LIMITED|PRIVATE LI|PVT LTD|PRIVATE|LIMITED|PVT|LTD)\b", "", d)
    d = re.sub(r"\d+$", "", d.strip())
    toks = [t for t in d.split() if not t.isdigit()]
    return " ".join(toks[:4]) or _clean(desc)[:20]


# ---------------------------------------------------------------- data access


def load_transactions() -> list[dict]:
    with TXN_CSV.open(encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    for r in rows:
        r["amount"] = float(r["amount"])
        r["merchant_key"] = merchant_key(r["description"])
    return rows


def group_merchants(rows: list[dict]) -> list[dict]:
    groups: dict[str, dict] = {}
    for r in rows:
        g = groups.setdefault(r["merchant_key"], {"merchant_key": r["merchant_key"], "count": 0, "total": 0.0, "sources": set(), "samples": [], "descriptions": [], "rule_category": set()})
        g["count"] += 1
        g["descriptions"].append(r["description"])
        g["total"] += r["amount"]
        g["sources"].add(r["source_id"])
        if r["description"] not in g["samples"] and len(g["samples"]) < 3:
            g["samples"].append(r["description"])
        if r.get("rule_category"):
            g["rule_category"].add(r["rule_category"] + (SEP + r["rule_subcategory"] if r.get("rule_subcategory") else ""))
    out = list(groups.values())
    out.sort(key=lambda g: -abs(g["total"]))
    return out


def load_suggestions() -> dict[str, dict]:
    if not SUGGESTIONS_CSV.exists():
        return {}
    with SUGGESTIONS_CSV.open(encoding="utf-8") as fh:
        return {r["merchant_key"]: r for r in csv.DictReader(fh)}


def taxonomy_paths() -> list[str]:
    import app as app_module

    paths = []
    for parent in app_module.PARENT_CATEGORIES:
        subs = getattr(app_module, "SUBCATEGORIES", {}).get(parent, [])
        paths.append(parent)  # parent only (no subcategory)
        paths.extend(parent + SEP + s for s in subs)
    return paths


def _split_path(path: str) -> tuple[str | None, str | None]:
    if not path:
        return None, None
    if SEP in path:
        cat, sub = path.split(SEP, 1)
        return cat.strip(), sub.strip()
    return path.strip(), None


# ---------------------------------------------------------------- commands


def cmd_merchants(_args) -> int:
    rows = load_transactions()
    groups = group_merchants(rows)
    print(f"{len(rows)} transactions, {len(groups)} merchant groups\n")
    print("count | total | sources | current_rule | merchant_key | samples")
    for g in groups:
        print(f"{g['count']:>4} | {g['total']:>11.2f} | {','.join(sorted(g['sources']))} | {','.join(sorted(g['rule_category'])) or '-'} | {g['merchant_key']} | {' || '.join(g['samples'])}")
    return 0


def cmd_build(_args) -> int:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    if XLSX.exists() and not getattr(_args, "force", False):
        print(f"{XLSX.name} already exists - it may hold your edits. Use 'migrate' to update it, or --force to overwrite.")
        return 1
    rows = load_transactions()
    groups = group_merchants(rows)
    sugg = load_suggestions()
    paths = taxonomy_paths()
    for s in sugg.values():  # any path Claude used that isn't in the app yet still shows in the dropdown
        p = (s.get("category") or "") + ((SEP + s["subcategory"]) if s.get("subcategory") else "")
        if p and p not in paths:
            paths.append(p)

    wb = Workbook()
    ws = wb.active
    ws.title = "Merchants"
    headers = ["merchant_key", "sample_description", "sources", "txn_count", "total_amount",
               "claude_category", "claude_flow", "claude_confidence", "claude_note",
               "FINAL_CATEGORY", "FINAL_FLOW", "YOUR_NOTE"]
    ws.append(headers)
    bold = Font(bold=True)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = bold
    ws.freeze_panes = "A2"
    low_fill = PatternFill("solid", fgColor="FFF4CC")
    edit_fill = PatternFill("solid", fgColor="E8F4EA")
    for g in groups:
        s = sugg.get(g["merchant_key"], {})
        path = (s.get("category") or "") + ((SEP + s["subcategory"]) if s.get("subcategory") else "")
        conf = float(s["confidence"]) if s.get("confidence") else ""
        ws.append([g["merchant_key"], g["samples"][0], ",".join(sorted(g["sources"])), g["count"], round(g["total"], 2),
                   path, s.get("flow_type", ""), conf, s.get("note", ""), path, s.get("flow_type", ""), ""])
        r = ws.max_row
        for c in (10, 11, 12):
            ws.cell(row=r, column=c).fill = edit_fill
        if conf != "" and conf < 0.7:
            for c in range(1, 10):
                ws.cell(row=r, column=c).fill = low_fill
    for i, w in enumerate([26, 48, 22, 9, 13, 34, 13, 10, 40, 34, 13, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = ws.dimensions

    _write_taxonomy_and_validation(wb, ws, paths)

    tt = wb.create_sheet("Transactions")
    tcols = ["merchant_key", "file", "date", "description", "amount", "rule_category", "rule_subcategory", "rule_flow_type"]
    tt.append(tcols)
    for c in range(1, len(tcols) + 1):
        tt.cell(row=1, column=c).font = bold
    for r in sorted(rows, key=lambda r: (r["merchant_key"], r["date"])):
        tt.append([r[c] for c in tcols])
    tt.freeze_panes = "A2"
    tt.auto_filter.ref = tt.dimensions
    for i, w in enumerate([26, 24, 12, 60, 12, 18, 24, 12], 1):
        tt.column_dimensions[get_column_letter(i)].width = w

    guide = wb.create_sheet("README", 0)
    guide.append(["How to review"])
    guide.cell(row=1, column=1).font = bold
    for line in [
        "1. Go to the Merchants sheet. Each row is one merchant (grouped across all statements).",
        "2. FINAL_CATEGORY is pre-filled with Claude's suggestion. Only change the ones that are wrong (dropdown: Parent › Subcategory).",
        "3. FINAL_FLOW: spend / income / refund / transfer / card_payment / fee / exclude. Change only if wrong.",
        "4. Need a subcategory that isn't in the dropdown? Add a row on the Taxonomy sheet as 'Parent › New Sub' and it appears in the dropdown.",
        "5. Yellow rows = Claude was unsure (confidence < 0.7). Green columns = yours to edit.",
        "6. Save the file, then run: python scripts/experiment_sheet.py apply",
    ]:
        guide.append([line])
    guide.column_dimensions["A"].width = 120
    wb.save(XLSX)
    missing = [g["merchant_key"] for g in groups if g["merchant_key"] not in sugg]
    print(f"wrote {XLSX} with {len(groups)} merchants ({len(missing)} without a suggestion)")
    for m in missing:
        print("  no suggestion:", m)
    return 0


def _write_taxonomy_and_validation(wb, ws, paths: list[str]) -> None:
    """(Re)create the Taxonomy sheet and the dropdown validations on Merchants!J:K.
    Excel drops openpyxl's validations when it re-saves, so migrate() calls this again."""
    from openpyxl.styles import Font
    from openpyxl.worksheet.datavalidation import DataValidation

    if "Taxonomy" in wb.sheetnames:
        del wb["Taxonomy"]
    tx = wb.create_sheet("Taxonomy")
    tx.append(["category_path (add new rows below to extend the dropdown)"])
    tx.cell(row=1, column=1).font = Font(bold=True)
    for p in paths:
        tx.append([p])
    tx.column_dimensions["A"].width = 44
    slack = 60
    tax_range = f"Taxonomy!$A$2:$A${1 + len(paths) + slack}"
    ws.data_validations.dataValidation = []
    dv_cat = DataValidation(type="list", formula1=tax_range, allow_blank=True, showErrorMessage=False)
    dv_flow = DataValidation(type="list", formula1='"' + ",".join(FLOW_CHOICES) + '"', allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_cat)
    ws.add_data_validation(dv_flow)
    last = max(ws.max_row, 2)
    dv_cat.add(f"J2:J{last}")
    dv_flow.add(f"K2:K{last}")


# Taxonomy changes to push through already-edited workbooks / suggestion files.
# old path -> new path, or a callable(merchant_key) -> new path. Anything else is left untouched.
MEAT_KEYS = {"LICIOUS", "MYCHICKENAN"}
DELIVERY_HINT = re.compile(r"ZOMATO|SWIGGY|ETERNAL|POWLE", re.I)
PATH_MIGRATIONS = [
    ("Food › Groceries / Quick Commerce",
     lambda key: "Groceries & Household › " + ("Meat & Fish" if key in MEAT_KEYS else "Groceries / Quick Commerce")),
    ("Food › Dining & Delivery",
     lambda key: "Food › " + ("Food Delivery" if DELIVERY_HINT.search(key or "") else "Dining")),
]


def migrate_path(path: str | None, key: str) -> str | None:
    if not path:
        return path
    for old, new in PATH_MIGRATIONS:
        if str(path).strip() == old:
            return new(key) if callable(new) else new
    return path


def cmd_migrate(_args) -> int:
    from openpyxl import load_workbook

    if not XLSX.exists():
        print(f"{XLSX} not found")
        return 1
    wb = load_workbook(XLSX)
    ws = wb["Merchants"]
    hdr = [c.value for c in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(hdr)}
    changed = 0
    present: list[str] = []
    for r in range(2, ws.max_row + 1):
        key = ws.cell(row=r, column=idx["merchant_key"]).value
        if not key:
            continue
        for col in ("claude_category", "FINAL_CATEGORY"):
            cell = ws.cell(row=r, column=idx[col])
            new = migrate_path(cell.value, key)
            if new != cell.value:
                cell.value = new
                changed += 1
            if cell.value and cell.value not in present:
                present.append(cell.value)
    old_tax = [c.value for c in wb["Taxonomy"]["A"][1:] if c.value] if "Taxonomy" in wb.sheetnames else []
    paths = taxonomy_paths()
    for extra in [migrate_path(p, "") for p in old_tax] + present:
        if extra and extra not in paths:
            paths.append(extra)
    _write_taxonomy_and_validation(wb, ws, paths)
    wb.save(XLSX)
    print(f"{XLSX.name}: {changed} category cells remapped; taxonomy list now {len(paths)} paths; dropdowns re-added")

    # keep the suggestions file consistent for any future rebuild
    if SUGGESTIONS_CSV.exists():
        with SUGGESTIONS_CSV.open(encoding="utf-8") as fh:
            rows = list(csv.DictReader(fh))
            fields = rows[0].keys() if rows else ["merchant_key", "category", "subcategory", "flow_type", "confidence", "note"]
        n = 0
        for row in rows:
            path = (row.get("category") or "") + ((SEP + row["subcategory"]) if row.get("subcategory") else "")
            new = migrate_path(path, row["merchant_key"])
            if new != path:
                cat, sub = _split_path(new)
                row["category"], row["subcategory"] = cat or "", sub or ""
                n += 1
        with SUGGESTIONS_CSV.open("w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=list(fields))
            w.writeheader()
            w.writerows(rows)
        print(f"{SUGGESTIONS_CSV.name}: {n} suggestions remapped")
    return 0


DEFAULT_CLASSIFICATION = {
    ("Food", None): "controllable",
    ("Groceries & Household", None): "baseline_variable",
    ("Home & Utilities", "Rent"): "fixed",
    ("Investments", None): "excluded",
    ("Travel", None): "one_off",
    ("Lifestyle", None): "controllable",
    ("Shopping", None): "controllable",
}


def classification_for(cat: str | None, sub: str | None, flow: str) -> str | None:
    if flow in ("transfer", "card_payment", "exclude", "income"):
        return "excluded"
    if not cat:
        return None
    return DEFAULT_CLASSIFICATION.get((cat, sub)) or DEFAULT_CLASSIFICATION.get((cat, None)) or "baseline_variable"


# Substrings that actually occur in the raw descriptions for the synthetic keys above.
KEY_PATTERNS = {
    "SWEEP IN FROM FD": "SWEEP TRF FROM",
    "SWEEP OUT TO FD": "SWEEP TRANSFER TO",
    "FD PREMATURE PROCEEDS": "FD PREMAT PROCEEDS",
    "EMPLOYER PAYROLL SALARY": "EMPLOYER PAYROLL",
    "HDFC CREDIT CARD PAYMENT (PG)": "PG OSHDFCCC",
    "ATM WITHDRAWAL": "ATL/",
    "NACH CREDIT (DIVIDEND/INTEREST)": "NACH-ECS-CR|NACH-10-CR",
    "DCC FEE": "DCC FEE",
    "INCOME TAX REFUND": "ITDTAX REFUND",
    "INDMONEY USD REMITTANCE": "TO INDMONE",
    "SAVINGS INTEREST": "INT.PD:",
    "UPI CREDIT ADJUSTMENT": "UPI_CRADJ",
    "CARD PAYMENT RECEIVED": "ONLINE PYMT RECD|PAYMENT RECEIVED",
    "IGST ON FEES": "IGST-VPS",
    "FCY MARKUP FEE": "FCY MARKUP FEE",
    "DCC MARKUP FEE": "DCC TRANSACTION",
    "PETRO SURCHARGE WAIVER": "PETRO SURCHARGE WAIVER",
}


def _common_prefix(strings: list[str], minimum: int = 5) -> str:
    ordered = sorted(strings, key=len)
    base = ordered[0]
    for n in ordered[1:]:
        while not n.startswith(base) and len(base) > minimum:
            base = base[:-1]
    return base.strip()


def _pattern_for(key: str, descriptions: list[str]) -> str | None:
    """A description_contains pattern (pipe = OR) that matches every description in the group."""
    def matches_all(pat: str) -> bool:
        alts = [a.strip().upper() for a in pat.split("|") if a.strip()]
        return all(any(a in d.upper() for a in alts) for d in descriptions)

    candidates: list[str] = []
    if key in KEY_PATTERNS:
        candidates.append(KEY_PATTERNS[key])
    upi_names = set()
    pci_names = set()
    for d in descriptions:
        u = d.upper()
        if "UPI/" in u:
            upi_names.add(_upi_merchant("UPI/" + d.split("UPI/", 1)[1]))
        m = re.match(r"^(PCI|PCD)/\d+/(.+?)/", d)
        if m:
            pci_names.add(m.group(2).split("*")[0].strip())
    if upi_names:
        ordered = sorted(upi_names, key=len)
        base = ordered[0]
        for n in ordered[1:]:
            while not n.startswith(base) and len(base) > 4:
                base = base[:-1]
        base = base.strip()
        candidates += ["UPI/" + base, base]
    if pci_names:
        ordered = sorted(pci_names, key=len)
        candidates.append("|".join(ordered) if len(ordered) > 1 else ordered[0])
        candidates.append(ordered[0])
    if key.startswith("NEFT ") or key.startswith("IMPS "):
        candidates.append(key.split(" ", 1)[1])
    # card statements: 'MERCHANTCITY' / 'MERCHANT* ref' -> longest common prefix of the raw text with city/ref/trailing digits removed
    card = []
    for d in descriptions:
        c = re.sub(r"\(Ref#.*?\)", "", d)
        c = _CITY_RE.sub("", c.strip())
        c = re.sub(r"\d+$", "", c.strip()).strip(" -*")
        card.append(c.upper())
    if card:
        candidates.append(_common_prefix(card))
    candidates.append(key)
    for c in candidates:
        if c and matches_all(c):
            return c
    return None


def _collisions(pattern: str, key: str, rows: list[dict]) -> int:
    alts = [a.strip().upper() for a in pattern.split("|") if a.strip()]
    return sum(1 for r in rows if r["merchant_key"] != key and any(a in r["description"].upper() for a in alts))


def cmd_apply(_args) -> int:
    from openpyxl import load_workbook

    import app as app_module

    rows = load_transactions()
    groups = {g["merchant_key"]: g for g in group_merchants(rows)}
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Merchants"]
    hdr = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(hdr)}
    decisions: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        key = row[idx["merchant_key"]]
        path = row[idx["FINAL_CATEGORY"]] or row[idx["claude_category"]] or ""
        flow = row[idx["FINAL_FLOW"]] or row[idx["claude_flow"]] or ""
        cat, sub = _split_path(str(path))
        changed = (row[idx["FINAL_CATEGORY"]] or "") != (row[idx["claude_category"]] or "") or (row[idx["FINAL_FLOW"]] or "") != (row[idx["claude_flow"]] or "")
        decisions[key] = {"category": cat, "subcategory": sub, "flow": str(flow), "note": row[idx["YOUR_NOTE"]] or "", "changed": changed}

    def decided(key: str) -> bool:
        d = decisions.get(key, {})
        return bool(d.get("category")) or d.get("flow", "") not in ("", "spend")

    known = getattr(app_module, "SUBCATEGORIES", {})
    used: dict[str, set] = defaultdict(set)
    for d in decisions.values():
        if d["category"]:
            used[d["category"]].add(d["subcategory"])
    new_parents = sorted(c for c in used if c not in app_module.PARENT_CATEGORIES)
    new_subs = {c: sorted(s for s in subs if s and s not in known.get(c, [])) for c, subs in used.items()}
    new_subs = {c: s for c, s in new_subs.items() if s}

    rules = []
    for key, d in decisions.items():
        g = groups.get(key)
        if not g or not decided(key):
            continue
        flow = "unknown" if d["flow"] == "exclude" else d["flow"]
        cls = classification_for(d["category"], d["subcategory"], d["flow"])
        pattern = _pattern_for(key, g["descriptions"])
        conf = 0.9 if g["count"] >= 3 else 0.85
        hits = _collisions(pattern, key, rows) if pattern else 0
        rules.append((key.title(), "description_contains", pattern, None, d["category"], d["subcategory"], cls, flow, conf, g["count"], round(g["total"], 2), hits))

    covered_txn = sum(g["count"] for k, g in groups.items() if decided(k))
    spend_total = sum(abs(r["amount"]) for r in rows if r["amount"] < 0) or 1
    covered_spend = sum(abs(r["amount"]) for r in rows if r["amount"] < 0 and decided(r["merchant_key"]))
    leftovers = [(k, g) for k, g in groups.items() if not decided(k)]

    lines = ["# Categorization experiment report", ""]
    lines += [f"- transactions: {len(rows)}; merchants: {len(groups)}; decisions changed by you: {sum(1 for d in decisions.values() if d['changed'])}",
              f"- coverage after experiment: {covered_txn}/{len(rows)} transactions ({covered_txn / len(rows):.0%}), {covered_spend / spend_total:.0%} of spend amount", ""]
    lines += ["## Taxonomy used", ""]
    for c in list(app_module.PARENT_CATEGORIES) + new_parents:
        subs = sorted(s for s in used.get(c, set()) if s)
        flag = "  <-- NEW PARENT" if c in new_parents else ""
        lines.append(f"- **{c}**{flag}: {', '.join(subs) or '(parent only)'}")
    lines += ["", "## Proposed SUBCATEGORIES additions", ""]
    if new_subs or new_parents:
        for c, subs in new_subs.items():
            lines.append(f"- {c}: + {', '.join(subs)}")
        for c in new_parents:
            lines.append(f"- NEW parent {c}: {', '.join(sorted(s for s in used[c] if s))}")
    else:
        lines.append("- none")
    lines += ["", "## Proposed default rules (seed_defaults tuple shape)", "", "```python"]
    for r in sorted(rules, key=lambda r: -abs(r[10])):
        name, mt, pat, _src, cat, sub, cls, flow, conf, n, tot, hits = r
        warn = ""
        if pat is None:
            warn = "  # !! NO PATTERN FOUND - write one by hand"
        elif hits:
            warn = f"  # !! also matches {hits} txns of other merchants"
        lines.append(f"    ({name!r}, {mt!r}, {pat!r}, None, {cat!r}, {sub!r}, {cls!r}, {flow!r}, {conf}),  # {n} txns, {tot}{warn}")
    lines += ["```", "", "## Still uncategorised", ""]
    for k, g in sorted(leftovers, key=lambda kg: -abs(kg[1]["total"])):
        lines.append(f"- {k}: {g['count']} txns, {g['total']:.2f} — {g['samples'][0]}")
    REPORT.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"wrote {REPORT}: {len(rules)} rules, {len(leftovers)} merchants uncategorised, coverage {covered_txn / len(rows):.0%}")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--dataset", help="suffix for the experiments/ files, e.g. 'diners' (default: the Kotak set)")
    sub = ap.add_subparsers(dest="cmd", required=True)
    sub.add_parser("merchants").set_defaults(fn=cmd_merchants)
    b = sub.add_parser("build")
    b.add_argument("--force", action="store_true", help="overwrite an existing (possibly edited) workbook")
    b.set_defaults(fn=cmd_build)
    sub.add_parser("apply").set_defaults(fn=cmd_apply)
    sub.add_parser("migrate").set_defaults(fn=cmd_migrate)
    args = ap.parse_args()
    use_dataset(args.dataset)
    return args.fn(args)


if __name__ == "__main__":
    raise SystemExit(main())
